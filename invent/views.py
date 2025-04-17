import csv, openpyxl, json, os, tempfile

from django.forms import forms
from django.shortcuts import render, redirect
from .models import *
from django.utils import timezone
from django.contrib import messages
from django.db.models import Sum, Q
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill
from django.template.loader import get_template, render_to_string
from weasyprint import HTML, CSS
from django.core.serializers.json import DjangoJSONEncoder
from django.core.mail import EmailMessage
from io import StringIO, BytesIO, TextIOWrapper
from django.conf import settings
from openpyxl import Workbook
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import get_user_model
from django.contrib.auth.forms import UserCreationForm
from django import forms
from .forms import CustomUserCreationForm

User = get_user_model()


@user_passes_test(lambda u: u.is_authenticated and u.is_super_admin)
def create_user_view(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/admin/')  # Or anywhere else you want
    else:
        form = CustomUserCreationForm()
    return render(request, 'invent/create_user.html', {'form': form})

def is_admin(user):
    return user.is_superuser or user.is_location_admin

# Create your views here.
@login_required
def index(request):
    return render(request, 'invent/index.html')

@login_required
def company_product(request):
    # Fetch all companies with their related products
    companies = Company.objects.prefetch_related('product_set').all()

    # Fetch staged entries for stock that is not yet added to inventory
    entries = StockEntry.objects.select_related('product', 'product__company').filter(staged=True)

    # Retrieve available locations (you may have a constant for location choices)
    locations = LOCATION_CHOICES  # Replace it with your location choices or dynamically fetched locations
    total_cost = sum(entry.total_cost for entry in entries)

    if request.method == "POST":
        if 'add_product' in request.POST:
            # Add new product to the staged list
            product_id = request.POST['product_id']
            quantity = int(request.POST['quantity'])
            product = Product.objects.get(id=product_id)
            location = request.POST.get('location')
            company = product.company


            StockEntry.objects.create(
                company=company,
                product=product,
                quantity=quantity,
                location=location,
                staged=True,
                created_by=request.user
            )
            messages.success(request, "Product added to stock.")
            return redirect('company_product')

        elif 'reset_list' in request.POST:
            # Reset staged products list
            StockEntry.objects.filter(staged=True).delete()
            messages.success(request, "List reset.")
            return redirect('company_product')

        elif 'delete_entry' in request.POST:
            # Delete specific staged product entry
            entry_id = request.POST['entry_id']
            StockEntry.objects.filter(id=entry_id, staged=True).delete()
            messages.success(request, "Entry deleted.")
            return redirect('company_product')

        elif 'add_to_stock' in request.POST:
            if not entries.exists():
                messages.warning(request, "No products to add.")
                return redirect('company_product')

            # Determine which location to use based on user role
            if request.user.is_superuser:
                location_value = request.POST.get('location')
            else:
                location_value = request.user.location  # Location admins are forced to their own location

            # Create a transaction for stock movement
            transaction = Transaction.objects.create()

            # Process each staged entry
            for entry in entries:
                # Update inventory based on the location
                inventory_item, created = Inventory.objects.get_or_create(
                    product=entry.product,
                    location=location_value,
                    defaults={'quantity': 0}
                )
                inventory_item.quantity += entry.quantity  # Add quantity to the inventory
                inventory_item.save()

                # Assign transaction and mark as committed
                entry.location = location_value
                entry.transaction = transaction
                entry.staged = False  # Mark entry as no longer staged
                entry.save()

            # Success message after updating inventory
            messages.success(request, f"Inventory updated. Transaction ID: {transaction.transaction_id}")
            return redirect('company_product')

    return render(request, 'invent/company_product.html', {
        'companies': companies,
        'entries': entries,
        'total_cost': total_cost,
        'locations': locations
    })


@login_required
def inventory_status(request):
    location = request.GET.get('location', 'All')
    search_query = request.GET.get('search', '')

    inventory = Inventory.objects.select_related('product')
    inventory_data = [{'product': item.product.name, 'quantity': item.quantity} for item in inventory]
    inventory_data = [{'product': item.product.name, 'quantity': item.quantity, 'location': item.location} for item in
                      inventory]

    if location != 'All':
        inventory = inventory.filter(location=location)

    if search_query:
        inventory = inventory.filter(product__name__icontains=search_query)

    total_quantity = inventory.aggregate(total=Sum('quantity'))['total'] or 0

    return render(request, 'invent/inventory_status.html', {
        'inventory': inventory,
        'locations': LOCATION_CHOICES,
        'selected_location': location,
        'search_query': search_query,
        'total_quantity': total_quantity,
        'inventory_data_json': json.dumps(inventory_data, cls=DjangoJSONEncoder)
    })

# CSV export view
@login_required
def export_inventory_csv(request):
    location = request.GET.get('location', 'All')
    search_query = request.GET.get('search', '')

    inventory = Inventory.objects.select_related('product')
    if location != 'All':
        inventory = inventory.filter(location=location)
    if search_query:
        inventory = inventory.filter(product__name__icontains=search_query)

    # Create a response with UTF-8 BOM
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="inventory_export.csv"'

    # Add UTF-8 BOM at the beginning
    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow(['Product Name', 'Location', 'Quantity'])
    for item in inventory:
        writer.writerow([item.product.name, item.location, item.quantity])

    return response


@login_required
def export_inventory_excel(request):
    location = request.GET.get('location', 'All')
    search_query = request.GET.get('search', '')

    inventory = Inventory.objects.select_related('product')
    if location != 'All':
        inventory = inventory.filter(location=location)
    if search_query:
        inventory = inventory.filter(product__name__icontains=search_query)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Report"

    headers = ['Product Name', 'Location', 'Quantity']
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    for item in inventory:
        ws.append([item.product.name, item.location, item.quantity])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventory.xlsx"'
    wb.save(response)
    return response

@login_required
def export_inventory_pdf(request):
    location = request.GET.get('location', 'All')
    search_query = request.GET.get('search', '')

    inventory = Inventory.objects.select_related('product')
    if location != 'All':
        inventory = inventory.filter(location=location)
    if search_query:
        inventory = inventory.filter(product__name__icontains=search_query)

    html_string = render_to_string('invent/inventory_pdf.html', {
        'inventory': inventory,
    })

    font_path = os.path.join(settings.STATIC_ROOT, 'fonts', 'NotoSansDevanagari-Regular.ttf')

    css = CSS(string=f"""
        @font-face {{
            font-family: NotoSansDevanagari;
            src: url("file://{font_path}") format("truetype");
        }}
        body {{
            font-family: 'NotoSansDevanagari', sans-serif;
        }}
    """)

    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf(stylesheets=[css])

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="inventory.pdf"'
    return response

@login_required
@csrf_exempt
def email_inventory_report(request):
    if request.method == "POST":
        email = request.POST['email']
        report_format = request.POST['format']
        inventory = Inventory.objects.select_related('product').all()

        file_data = None
        mime_type = ""
        filename = ""

        if report_format == "csv":
            buffer = StringIO()
            buffer.write('\ufeff')  # Add BOM for UTF-8
            writer = csv.writer(buffer)
            writer.writerow(["Product", "Location", "Quantity"])
            for item in inventory:
                writer.writerow([item.product.name, item.location, item.quantity])
            file_data = buffer.getvalue().encode("utf-8")
            mime_type = "text/csv"
            filename = "inventory_report.csv"

        elif report_format == "excel":
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Product", "Location", "Quantity"])
            for item in inventory:
                sheet.append([item.product.name, item.location, item.quantity])
            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)
            file_data = buffer.read()
            mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = "inventory_report.xlsx"

        elif report_format == "pdf":
            html_string = render_to_string("invent/inventory_pdf.html", {'inventory': inventory})
            pdf_file = BytesIO()
            HTML(string=html_string).write_pdf(pdf_file)
            pdf_file.seek(0)
            file_data = pdf_file.read()
            mime_type = "application/pdf"
            filename = "inventory_report.pdf"

        # Send email
        if file_data:
            email_msg = EmailMessage(
                subject="Inventory Report",
                body="Please find attached your inventory report.",
                from_email="your_email@example.com",
                to=[email],
            )
            email_msg.attach(filename, file_data, mime_type)
            email_msg.send()

        return redirect("inventory_status")

    return HttpResponse("Invalid method", status=405)

@login_required
@csrf_exempt
def available_products(request):
    if request.method == "POST":
        data = json.loads(request.body)
        location = data.get("location")

        inventory_items = Inventory.objects.filter(location=location)
        product_list = []

        for inv in inventory_items:
            product_list.append({
                "name": inv.product.name,
                "cost": float(inv.product.cost),
                "quantity": inv.quantity,
            })

        return JsonResponse({"products": product_list})

    return JsonResponse({"error": "Invalid request"}, status=400)

@login_required
@csrf_exempt
def stock_transfer_view(request):
    if request.method == "POST":
        data = json.loads(request.body)
        from_location = data.get("from_location")
        to_location = data.get("to_location")
        items = data.get("items", [])

        if from_location == to_location:
            return JsonResponse({"success": False, "error": "From and To locations must be different."})

        for item in items:
            product_name = item.get("product")
            quantity = int(item.get("quantity"))

            try:
                product = Product.objects.get(name=product_name)
            except Product.DoesNotExist:
                return JsonResponse({"success": False, "error": f"Product '{product_name}' not found."})

            # Check stock at from_location
            from_inv, _ = Inventory.objects.get_or_create(product=product, location=from_location)
            if from_inv.quantity < quantity:
                return JsonResponse({"success": False, "error": f"Not enough stock of '{product_name}' at {from_location}."})

            # Update inventory
            from_inv.quantity -= quantity
            from_inv.save()

            to_inv, _ = Inventory.objects.get_or_create(product=product, location=to_location)
            to_inv.quantity += quantity
            to_inv.save()

            # Create StockTransfer record
            StockTransfer.objects.create(
                transaction_id=f"TXN{int(timezone.now().timestamp())}",
                product=product,
                from_location=from_location,
                to_location=to_location,
                quantity=quantity
            )

        return JsonResponse({"success": True, "message": "Stock transferred successfully."})

    return render(request, "invent/stock_transfer.html")

@login_required
def transaction_log_view(request):
    date = request.GET.get('date')
    location = request.GET.get('location')
    transaction_id = request.GET.get('transaction_id')
    product_name = request.GET.get('product_name')

    stock_entries = StockEntry.objects.select_related('product__company', 'transaction')
    transfers = StockTransfer.objects.select_related('product')

    stock_entries_data = []
    stock_transfers_data = []

    # Apply filters
    if date:
        stock_entries = stock_entries.filter(date=date)
        transfers = transfers.filter(timestamp__date=date)

    if location:
        stock_entries = stock_entries.filter(location__icontains=location)
        transfers = transfers.filter(Q(from_location__icontains=location) | Q(to_location__icontains=location))

    if transaction_id:
        # 'iexact' for exact match (case-insensitive) on the `transaction_id` field.
        stock_entries = stock_entries.filter(transaction__transaction_id__iexact=transaction_id)
        transfers = transfers.filter(transaction_id__iexact=transaction_id)

    if product_name:
        # Correct filter for ForeignKey field (targeting the 'name' of 'product')
        stock_entries = stock_entries.filter(product__name__icontains=product_name)
        transfers = transfers.filter(product__name__icontains=product_name)

    # Prepare the stock entries data
    for entry in stock_entries:
        # Safely get the first matching inventory location
        inventory = Inventory.objects.filter(product=entry.product).first()
        location = inventory.location if inventory else "Unknown"

        stock_entries_data.append({
            "transactionId": entry.transaction.transaction_id if entry.transaction else "Staged",
            "date": entry.date.isoformat(),
            "company": entry.product.company.name,
            "product": entry.product.name,
            "quantity": entry.quantity,
            "totalCost": float(entry.total_cost),
            "location": location  # Use fetched location
        })

    # Prepare the stock transfers data
    stock_transfers_data = [
        {
            "transactionId": transfer.transaction_id,
            "date": transfer.timestamp.date().isoformat(),
            "fromLocation": transfer.from_location,
            "toLocation": transfer.to_location,
            "product": transfer.product.name,
            "quantity": transfer.quantity
        } for transfer in transfers
    ]

    return render(request, 'invent/transaction_log.html', {
        'stock_entries': stock_entries_data,
        'stock_transfers': stock_transfers_data,
        'locations': [choice[0] for choice in LOCATION_CHOICES],
    })


def export_stock_entries_csv(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="stock_entries.csv"'
    response.write('\ufeff'.encode('utf8'))  # UTF-8 BOM for Excel

    writer = csv.writer(response)
    writer.writerow(['Transaction ID', 'Date', 'Company', 'Product', 'Quantity', 'Cost', 'Location'])

    for entry in StockEntry.objects.all():
        writer.writerow([
            entry.transaction.transaction_id if entry.transaction else "Staged",
            entry.date.strftime("%d-%m-%Y"),
            entry.company.name,
            entry.product.name,
            entry.quantity,
            float(entry.total_cost),
            entry.location
        ])
    return response


def export_stock_transfers_csv(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="stock_transfers.csv"'
    response.write('\ufeff'.encode('utf8'))

    writer = csv.writer(response)
    writer.writerow(['Transaction ID', 'Date', 'From Location', 'To Location', 'Product', 'Quantity'])

    for transfer in StockTransfer.objects.all():
        writer.writerow([
            transfer.transaction_id,
            transfer.timestamp.strftime("%d-%m-%Y"),
            transfer.from_location,
            transfer.to_location,
            transfer.product,
            transfer.quantity
        ])
    return response


def export_stock_entries_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Entries"

    ws.append(['Transaction ID', 'Date', 'Company', 'Product', 'Quantity', 'Cost', 'Location'])

    for entry in StockEntry.objects.all():
        ws.append([
            entry.transaction.transaction_id if entry.transaction else "Staged",
            entry.date.strftime("%d-%m-%Y"),
            entry.company.name,
            entry.product.name,
            entry.quantity,
            float(entry.total_cost),
            entry.location
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="stock_entries.xlsx"'
    wb.save(response)
    return response


def export_stock_transfers_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Transfers"

    ws.append(['Transaction ID', 'Date', 'From Location', 'To Location', 'Product', 'Quantity'])

    for transfer in StockTransfer.objects.all():
        ws.append([
            transfer.transaction_id,
            transfer.timestamp.strftime("%Y-%m-%d"),
            transfer.from_location,
            transfer.to_location,
            transfer.product.name,
            transfer.quantity
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="stock_transfers.xlsx"'
    wb.save(response)
    return response



def export_stock_entries_pdf(request):
    location = request.GET.get('location', 'All')
    search_query = request.GET.get('search', '')

    # Filter the stock entries based on location and search query
    stock_entries = StockEntry.objects.select_related('company', 'product')
    if location != 'All':
        stock_entries = stock_entries.filter(location=location)
    if search_query:
        stock_entries = stock_entries.filter(product__name__icontains=search_query)

    # Render the HTML template
    html_string = render_to_string('invent/stock_entries_pdf.html', {
        'stock_entries': stock_entries,
    })

    # Path to the custom font
    font_path = os.path.join(settings.STATIC_ROOT, 'fonts', 'NotoSansDevanagari-Regular.ttf')

    # Define custom CSS for the PDF
    css = CSS(string=f"""
        @font-face {{
            font-family: NotoSansDevanagari;
            src: url("file://{font_path}") format("truetype");
        }}
        body {{
            font-family: 'NotoSansDevanagari', sans-serif;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
        }}
        th, td {{
            padding: 8px;
            text-align: left;
            border: 1px solid #ddd;
        }}
        th {{
            background-color: #f2f2f2;
        }}
    """)

    # Generate the PDF from the HTML string and CSS
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf(stylesheets=[css])

    # Return the PDF as an HTTP response
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="stock_entries.pdf"'
    return response


def export_stock_transfers_pdf(request):
    location_from = request.GET.get('from_location', 'All')
    location_to = request.GET.get('to_location', 'All')
    search_query = request.GET.get('search', '')

    # Filter the stock transfers based on location and search query
    stock_transfers = StockTransfer.objects.select_related('product')
    if location_from != 'All':
        stock_transfers = stock_transfers.filter(from_location=location_from)
    if location_to != 'All':
        stock_transfers = stock_transfers.filter(to_location=location_to)
    if search_query:
        stock_transfers = stock_transfers.filter(product__name__icontains=search_query)

    # Render the HTML template
    html_string = render_to_string('invent/stock_transfers_pdf.html', {
        'stock_transfers': stock_transfers,
    })

    # Path to the custom font
    font_path = os.path.join(settings.STATIC_ROOT, 'fonts', 'NotoSansDevanagari-Regular.ttf')

    # Define custom CSS for the PDF
    css = CSS(string=f"""
        @font-face {{
            font-family: NotoSansDevanagari;
            src: url("file://{font_path}") format("truetype");
        }}
        body {{
            font-family: 'NotoSansDevanagari', sans-serif;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
        }}
        th, td {{
            padding: 8px;
            text-align: left;
            border: 1px solid #ddd;
        }}
        th {{
            background-color: #f2f2f2;
        }}
    """)

    # Generate the PDF from the HTML string and CSS
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf(stylesheets=[css])

    # Return the PDF as an HTTP response
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="stock_transfers.pdf"'
    return response
